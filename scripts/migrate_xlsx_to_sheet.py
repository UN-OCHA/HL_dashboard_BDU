#!/usr/bin/env python3
"""migrate_xlsx_to_sheet.py — one-off data migration.

Reads `HC trends report 2025.xlsx` and writes 10 CSVs under
`data/initial/` that match the tab structure of the production
Google Sheet. Paste each CSV's contents into the corresponding tab.

Usage:
    python3 scripts/migrate_xlsx_to_sheet.py \\
        --xlsx "../data/HC trends report 2025.xlsx" \\
        --out  "data/initial"

Requires: openpyxl  (pip install openpyxl)

The script is idempotent and safe to re-run.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional


try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.stderr.write("ERROR: openpyxl is required. Run: pip install openpyxl\n")
    sys.exit(2)


# ── Country name → ISO3 map for the subset of countries in the snapshot ──
# Keep as exact matches for the PPT's Slide-5 country column. Extend as needed.
NAME_TO_ISO3 = {
    # Duty-station countries (Slide 5)
    "afghanistan": "AFG", "burkina faso": "BFA", "cameroon": "CMR",
    "central african republic": "CAF", "chad": "TCD", "colombia": "COL",
    "democratic republic of the congo": "COD", "drc": "COD",
    "eritrea": "ERI", "ethiopia": "ETH", "haiti": "HTI",
    "iraq": "IRQ", "lebanon": "LBN", "mali": "MLI",
    "mozambique": "MOZ", "myanmar": "MMR", "niger": "NER", "nigeria": "NGA",
    "occupied palestinian territory": "PSE", "opt": "PSE", "opt gaza": "PSE",
    "state of palestine": "PSE", "palestine": "PSE",
    "pakistan": "PAK", "somalia": "SOM", "south sudan": "SSD",
    "sudan": "SDN", "sudan crisis": "SDN",
    "syrian arab republic": "SYR", "syria": "SYR",
    "ukraine": "UKR",
    "venezuela (bolivarian republic of)": "VEN", "venezuela": "VEN",
    "yemen": "YEM", "zimbabwe": "ZWE",
    # Nationalities that commonly appear as HCs
    "algeria": "DZA", "argentina": "ARG", "australia": "AUS", "austria": "AUT",
    "bangladesh": "BGD", "belgium": "BEL", "benin": "BEN", "brazil": "BRA",
    "bulgaria": "BGR", "burundi": "BDI", "canada": "CAN", "chile": "CHL",
    "china": "CHN", "côte d'ivoire": "CIV", "cote d'ivoire": "CIV",
    "croatia": "HRV", "denmark": "DNK", "egypt": "EGY", "finland": "FIN",
    "france": "FRA", "germany": "DEU", "ghana": "GHA", "greece": "GRC",
    "india": "IND", "indonesia": "IDN", "iran (islamic republic of)": "IRN",
    "iran": "IRN", "ireland": "IRL", "italy": "ITA", "japan": "JPN",
    "jordan": "JOR", "kenya": "KEN", "liberia": "LBR", "malaysia": "MYS",
    "morocco": "MAR", "nepal": "NPL", "netherlands": "NLD", "new zealand": "NZL",
    "norway": "NOR", "peru": "PER", "philippines": "PHL", "poland": "POL",
    "portugal": "PRT", "romania": "ROU",
    "russian federation": "RUS", "russia": "RUS",
    "rwanda": "RWA", "senegal": "SEN", "sierra leone": "SLE",
    "south africa": "ZAF", "spain": "ESP", "sri lanka": "LKA",
    "sweden": "SWE", "switzerland": "CHE", "tanzania": "TZA",
    "united republic of tanzania": "TZA",
    "thailand": "THA", "tunisia": "TUN", "turkey": "TUR", "türkiye": "TUR",
    "uganda": "UGA", "united kingdom": "GBR", "uk": "GBR",
    "united states of america": "USA", "united states": "USA", "usa": "USA",
    "uruguay": "URY", "zambia": "ZMB",
}


def norm(s: Optional[str]) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def iso3_for(country: str) -> str:
    return NAME_TO_ISO3.get(norm(country), "")


def excel_date_to_iso(v) -> str:
    """Best-effort conversion of whatever openpyxl gives us into ISO date."""
    if v in (None, ""):
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, (int, float)):
        # Excel 1900 serial → date
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(v))).date().isoformat()
    s = str(v).strip()
    # Try a few common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s


def sheet_records(ws, header_signals: Optional[List[str]] = None) -> List[Dict[str, object]]:
    """Read a worksheet as a list of dicts keyed by lowercased headers.

    header_signals: one or more lowercased strings that must appear in the
    header row. If provided, the first row containing any of them is treated
    as the header. Otherwise the very first row is used (legacy behaviour).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = 0
    if header_signals:
        signals = [s.lower() for s in header_signals]
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if any(sig in cells for sig in signals):
                header_idx = i
                break

    headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1:]:
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    return out


def write_csv(path: Path, header: List[str], rows: Iterable[List[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            w.writerow(r)
    print(f"  wrote {path.relative_to(path.parent.parent)}  ({path.stat().st_size:,} B)")


# ── Per-tab builders ─────────────────────────────────────────────────────────

# Exact 29-row roster copied from PPT Slide 5. Same columns as the
# Google Sheet `leaders` tab. Valijon edits in the Sheet; this seeds
# the starter CSV so the dashboard matches the published snapshot.
PPT_LEADERS = [
    # country, duty_station, name, position
    ("Afghanistan",                   "Kabul",        "Indrika Ratwatte",              "DSRSG/RC/HC"),
    ("Burkina Faso",                  "Ouagadougou",  "Maurice Azonnankpo",            "RC/HC OiC"),
    ("Cameroon",                      "Yaoundé",      "Issa Sanogo",                   "RC/HC"),
    ("Central African Republic",      "Bangui",       "Mohamed Ag Ayoya",              "DSRSG/RC/HC"),
    ("Chad",                          "N'Djamena",    "François Batalingaya",          "RC/HC"),
    ("Colombia",                      "Bogota",       "Mireia Villar Forner",          "RC/HC"),
    ("Democratic Republic of the Congo", "Kinshasa",  "Bruno Lemarquis",               "DSRSG/RC/HC"),
    ("Eritrea",                       "Asmara",       "Nahla Valji",                   "RC/HC"),
    ("Ethiopia",                      "Addis Ababa",  "Aboubacar Kampo",               "RC/HC OiC"),
    ("Haiti",                         "Port-au-Prince", "Nicole Kouassi",              "DSRSG/RC/HC"),
    ("Iraq",                          "Baghdad",      "Ghulam Isaczai",                "DSRSG/RC/HC"),
    ("Lebanon",                       "Beirut",       "Imran Riza",                    "DSC/RC/HC"),
    ("Mali",                          "Bamako",       "Hanaa Singer",                  "RC/HC"),
    ("Mozambique",                    "Maputo",       "Catherine Sozi",                "RC/HC"),
    ("Myanmar",                       "Yangon",       "Gwyn Lewis",                    "RC/HC a.i."),
    ("Niger",                         "Niamey",       "Mama Keita",                    "RC/HC"),
    ("Nigeria",                       "Abuja",        "Mohamed Fall",                  "RC/HC"),
    ("OPT",                           "Jerusalem",    "Ramiz Alakbarov",               "DSC/RC/HC"),
    ("Pakistan",                      "Islamabad",    "Mohamed Yahya",                 "RC/HC"),
    ("Somalia",                       "Mogadishu",    "George Conway",                 "DSRSG/RC/HC"),
    ("South Sudan",                   "Juba",         "Anita Kiki Gbeho",              "DSRSG/RC/HC"),
    ("Sudan",                         "Port Sudan",   "Denise Brown",                  "RC/HC"),
    ("Syrian Arab Republic",          "Damascus",     "Nathalie Fustier",              "RC/HC a.i."),
    ("Ukraine",                       "Kyiv",         "Matthias Schmale",              "RC/HC"),
    ("Venezuela",                     "Caracas",      "Gianluca Rampolla del Tindaro", "RC/HC"),
    ("Yemen",                         "Sana'a",       "Julien Harneis",                "RC/HC"),
    ("Zimbabwe",                      "Harare",       "Edward Kallon",                 "RC/HC"),
    ("OPT",                           "Gaza",         "Suzanna Tkalec",                "Deputy HC"),
    ("Sudan Crisis",                  "Tawila",       "Rosaria Bruno",                 "Deputy HC"),
]


def build_leaders(_active: List[Dict]) -> tuple[List[str], List[List]]:
    """Seed leaders tab verbatim from PPT Slide 5 (29 rows).

    Columns kept minimal — matches the PPT table exactly. Extra columns
    (hat3, gender, nationality, agency, eod, eoa) are blank so the Sheet
    remains the source of truth for those details.
    """
    header = [
        "country", "duty_station", "name", "position",
        "hat3", "gender", "nationality", "weog", "agency", "eod", "eoa"
    ]
    rows = [
        [c, d, n, p, "", "", "", "", "", "", ""]
        for (c, d, n, p) in PPT_LEADERS
    ]
    return header, rows


def _build_leaders_from_xlsx(active: List[Dict]) -> tuple[List[str], List[List]]:
    """Current roster from the 'HC trends report 01 2025 - LF' sheet.

    NOTE: the source xlsx does not carry a 'country of duty station' column.
    As a starter, we populate `country` and `duty_station` with `nationality`
    so the dashboard renders. Valijon should override these in the Google
    Sheet with the actual duty-station country for each leader.
    """
    header = [
        "country", "duty_station", "name", "position", "hat3",
        "gender", "nationality", "weog", "agency", "eod", "eoa"
    ]
    rows = []
    for r in active:
        name = r.get("full name") or r.get("full_name") or ""
        pos  = (r.get("position title") or r.get("position_title") or "").strip()
        hat  = (r.get("hat 3") or r.get("hat3") or "").strip()
        # Combine Position Title + Hat 3 into a full role label.
        # Avoid doubling when Hat already appears in the position string
        # (e.g. "DSRSG/RC/HC" already contains "HC").
        if hat and hat.upper() not in pos.upper():
            full_pos = (pos + "/" + hat).strip("/")
        else:
            full_pos = pos or hat
        nat  = r.get("nationality") or ""
        # No duty-station country in source → fall back to nationality.
        country = r.get("country") or r.get("duty station") or r.get("duty_station") or nat
        duty    = r.get("duty station") or r.get("duty_station") or country
        rows.append([
            country, duty, name, full_pos, hat,
            r.get("gender") or "",
            nat,
            r.get("weog/non-weog") or r.get("weog/non_weog") or "",
            r.get("agency of origin short name") or r.get("agency") or "",
            excel_date_to_iso(r.get("entrance on duty") or r.get("eod")),
            excel_date_to_iso(r.get("end of assignment") or r.get("eoa")),
        ])
    return header, rows


def build_contacts(active: List[Dict]) -> tuple[List[str], List[List]]:
    """Contact directory. The source Excel does NOT have emails/phones —
    those come from a separate roster the user will paste into the Sheet.
    We output the names + positions now as a scaffold."""
    header = [
        "country", "name", "position", "email", "phone",
        "pa_name", "pa_phone", "ea_name", "ea_phone"
    ]
    rows = []
    for r in active:
        country = (r.get("country") or r.get("duty station") or r.get("duty_station")
                   or r.get("nationality") or "")
        name = r.get("full name") or ""
        pos  = (r.get("position title") or "").strip()
        hat  = (r.get("hat 3") or "").strip()
        if hat and hat.upper() not in pos.upper():
            pos = (pos + "/" + hat).strip("/")
        rows.append([country, name, pos, "", "", "", "", "", ""])
    return header, rows


# Canonical duty-station roster from PPT slide 5 (Feb 2026 snapshot).
# Columns: iso3, country (display name), primary_role, has_dhc.
#   has_dhc = "yes" means the country has an additional Deputy HC on top
#   of the primary role — rendered on the map with a diagonal stripe
#   pattern overlay ("+ DHC" in the legend).
# Role values match the 4 PPT categories:
#   • RC/HC           (covers "RC/HC", "RC/HC OiC", "RC/HC a.i.")
#   • DSRSG/RC/HC
#   • DSC/RC/HC
#   • Deputy HC        (standalone DHC post, no other role)
PPT_MAP_COUNTRIES = [
    # iso3,  country,                          primary_role,     has_dhc
    ("AFG",  "Afghanistan",                    "DSRSG/RC/HC",    "no"),
    ("BFA",  "Burkina Faso",                   "RC/HC",          "no"),
    ("CMR",  "Cameroon",                       "RC/HC",          "no"),
    ("CAF",  "Central African Republic",       "DSRSG/RC/HC",    "no"),
    ("TCD",  "Chad",                           "RC/HC",          "no"),
    ("COL",  "Colombia",                       "RC/HC",          "no"),
    ("COD",  "Democratic Republic of the Congo","DSRSG/RC/HC",   "no"),
    ("ERI",  "Eritrea",                        "RC/HC",          "no"),
    ("ETH",  "Ethiopia",                       "RC/HC",          "no"),
    ("HTI",  "Haiti",                          "DSRSG/RC/HC",    "no"),
    ("IRQ",  "Iraq",                           "DSRSG/RC/HC",    "no"),
    ("LBN",  "Lebanon",                        "DSC/RC/HC",      "no"),
    ("MLI",  "Mali",                           "RC/HC",          "no"),
    ("MOZ",  "Mozambique",                     "RC/HC",          "no"),
    ("MMR",  "Myanmar",                        "RC/HC",          "no"),
    ("NER",  "Niger",                          "RC/HC",          "no"),
    ("NGA",  "Nigeria",                        "RC/HC",          "no"),
    # OPT has both a DSC/RC/HC (Jerusalem) and a Deputy HC (Gaza).
    ("PSE",  "OPT",                            "DSC/RC/HC",      "yes"),
    ("PAK",  "Pakistan",                       "RC/HC",          "no"),
    ("SOM",  "Somalia",                        "DSRSG/RC/HC",    "no"),
    ("SSD",  "South Sudan",                    "DSRSG/RC/HC",    "no"),
    # Sudan has both an RC/HC (Port Sudan) and a Deputy HC (Tawila).
    ("SDN",  "Sudan",                          "RC/HC",          "yes"),
    ("SYR",  "Syrian Arab Republic",           "RC/HC",          "no"),
    ("UKR",  "Ukraine",                        "RC/HC",          "no"),
    ("VEN",  "Venezuela",                      "RC/HC",          "no"),
    ("YEM",  "Yemen",                          "RC/HC",          "no"),
    ("ZWE",  "Zimbabwe",                       "RC/HC",          "no"),
]


def build_map_countries(_active: List[Dict]) -> tuple[List[str], List[List]]:
    """Seed map_countries from the PPT Slide 5 roster verbatim.

    The live dashboard expects this data to come from the Google Sheet —
    Valijon can edit any row there without code changes. This starter CSV
    mirrors the PPT so the first render matches the printed snapshot.
    """
    header = ["iso3", "country", "primary_role", "has_dhc"]
    return header, [list(r) for r in PPT_MAP_COUNTRIES]


# Page 3 chart values are copied verbatim from the PPT — see PPT_* constants
# below. The old xlsx-derived roles/agency/country-by-grade/gender-by-grade
# builders were removed because the PPT data is the source of truth.


# Exact figures copied from the PPT "Humanitarian Leadership Snapshot".
# Page 3's two grade-based bar charts are a direct mirror of PPT slides
# 3 (Country of Origin) and 4 (Grade and Gender). The source Excel does
# not carry a grade column, so we ship these values as the starter CSVs
# and let Valijon edit them in the Google Sheet when figures change.
PPT_COUNTRY_BY_GRADE = [  # grade, weog, non_weog  — from PPT slide 3
    ["ASG", 6, 8],
    ["D2",  3, 8],
    ["D1",  2, 0],
]
PPT_GENDER_BY_GRADE = [   # grade, female, male    — from PPT slide 4
    ["ASG", 4, 10],
    ["D2",  6, 5],
    ["D1",  2, 0],
]
PPT_ROLES_DONUT = [       # label, value          — from PPT slide 3 donut
    ["RC/HC",        16],
    ["DSRSG/RC/HC",   7],
    ["DSC/RC/HC",     2],
    ["Deputy HC",     2],
]
PPT_AGENCY_DONUT = [      # label, value          — from PPT slide 3 donut
    ["Other",   9],
    ["UNDP",    6],
    ["UNICEF",  5],
    ["OCHA",    4],
    ["WFP",     3],
]


def build_country_by_grade(_active: List[Dict]) -> tuple[List[str], List[List]]:
    return ["grade", "weog", "non_weog"], [list(r) for r in PPT_COUNTRY_BY_GRADE]


def build_gender_by_grade(_active: List[Dict]) -> tuple[List[str], List[List]]:
    return ["grade", "female", "male"], [list(r) for r in PPT_GENDER_BY_GRADE]


def build_roles_donut_from_ppt(_active: List[Dict]) -> tuple[List[str], List[List]]:
    return ["label", "value"], [list(r) for r in PPT_ROLES_DONUT]


def build_agency_donut_from_ppt(_active: List[Dict]) -> tuple[List[str], List[List]]:
    return ["label", "value"], [list(r) for r in PPT_AGENCY_DONUT]


def build_gender_trends(gender_sheet: List[Dict]) -> tuple[List[str], List[List]]:
    header = ["year", "female_pct", "male_pct"]
    rows: List[List] = []
    for r in gender_sheet:
        year = r.get("year") or r.get("row labels")
        if not isinstance(year, (int, float)):
            try:
                year = int(str(year).strip())
            except Exception:
                continue
        total = (r.get("female") or 0) + (r.get("male") or 0)
        if total <= 0:
            continue
        f_pct = round((r.get("female") or 0) / total * 100, 1)
        m_pct = round((r.get("male") or 0) / total * 100, 1)
        rows.append([int(year), f_pct, m_pct])
    rows.sort(key=lambda x: x[0])
    return header, rows


def build_region_trends(region_sheet: List[Dict]) -> tuple[List[str], List[List]]:
    header = ["year", "africa_pct", "apac_pct", "eeur_pct", "lac_pct", "weog_pct"]
    rows: List[List] = []
    for r in region_sheet:
        year = r.get("year") or r.get("row labels")
        if not isinstance(year, (int, float)):
            try:
                year = int(str(year).strip())
            except Exception:
                continue
        def g(k: str) -> float:
            v = r.get(k)
            return float(v) if isinstance(v, (int, float)) else 0.0
        total = g("africa") + g("asia and the pacific") + g("eastern europe") + \
                g("latin america and the caribbean") + g("weog")
        if total <= 0:
            continue
        rows.append([
            int(year),
            round(g("africa") / total * 100, 1),
            round(g("asia and the pacific") / total * 100, 1),
            round(g("eastern europe") / total * 100, 1),
            round(g("latin america and the caribbean") / total * 100, 1),
            round(g("weog") / total * 100, 1),
        ])
    rows.sort(key=lambda x: x[0])
    return header, rows


def build_meta(active: List[Dict]) -> tuple[List[str], List[List]]:
    """Boilerplate meta the user can edit in the Sheet."""
    n = len(active)
    females = sum(1 for r in active if str(r.get("gender", "")).lower().startswith("f"))
    non_weog = sum(1 for r in active if str(r.get("weog/non-weog", "")).lower().startswith("non"))
    deputies = sum(
        1 for r in active
        if str(r.get("hat 3", "")).strip().upper() == "DHC"
        or "deputy" in str(r.get("position title", "")).lower()
    )
    # Count distinct duty-station (or nationality-fallback) countries that
    # we could resolve to ISO3 — keeps the KPI consistent with the map.
    countries = set()
    for r in active:
        c = (r.get("country") or r.get("duty station") or r.get("duty_station")
             or r.get("nationality") or "")
        iso = iso3_for(c)
        if iso:
            countries.add(iso)

    def pct(x, total):
        return round((x / total) * 100) if total else 0

    header = ["key", "value"]
    today = datetime.today().strftime("%d %B %Y")
    month = datetime.today().strftime("%B %Y")

    # ── Verbatim PPT text for each editable text slot ────────────────

    # Slide 1 — OVERVIEW + LEADERSHIP ON THE MOVE sidebar.
    overview = (
        "There are currently **27 humanitarian leaders** coordinating humanitarian "
        "response in **25 countries**.\n\n"
        "Of this leadership cohort, 7 are Deputy Special Representatives of the "
        "Secretary-General / Resident Coordinators / Humanitarian Coordinators "
        "(DSRSG/RC/HC) and 2 Deputy Special Coordinators / RC/HCs (DSC/RC/HC), "
        "16 RC/HCs, and 2 Deputy HCs (DHCs).\n\n"
        "Over **44 per cent** of the cohort is female, while **60 per cent** "
        "originate from countries historically under-represented."
    )
    leadership_on_move = (
        "**Regional:** Mr. John Ging has been designated as Senior Advisor on the "
        "Middle East and North Africa (MENA)."
    )

    # Slide 2 — monthly highlight card (EDG endorsement paragraph).
    highlight = (
        "The IASC Emergency Directors Group (EDG) recently endorsed the HC "
        "Leadership Profile, which sets out a clear framework for the "
        "**values, attributes, competencies, and commitments** expected of "
        "effective humanitarian leaders. The Profile complements the RC "
        "Leadership Profile and HC Terms of Reference, guides HC designations, "
        "selection into the HC Pool and Talent Pipeline, and performance "
        "management. While it reflects the full range of qualities expected "
        "of HCs, it recognizes that specific technical skills, experience, "
        "and contextual knowledge may differ depending on the operational "
        "environment."
    )

    # Slide 3 — observations sidebar under the characteristic charts.
    note_characteristics = (
        "Leadership positions are largely concentrated at the **RC/HC level**, "
        "representing around 60 per cent of the total, followed by DSRSG/RC/HC "
        "positions at 26 per cent, with Deputy HC and other senior roles "
        "accounting for the remaining 15 per cent.\n\n"
        "The current leadership cadre is drawn from a range of agencies: "
        "22 per cent from UNDP, 19 per cent from UNICEF, 15 per cent from OCHA, "
        "and approximately 33 per cent from nine other agencies, each "
        "represented by fewer than three individuals.\n\n"
        "Representation by country of origin varies across grades: at the ASG "
        "level, representation is relatively balanced, while Non-WEOG "
        "representation is stronger at D2 and more limited at the D1 level."
    )

    # Slide 4 — observations sidebar under the trend charts.
    note_trends = (
        "Gender distribution is different across leadership grades. At the "
        "**ASG level**, women account for around 30 per cent of the total. "
        "At the **D2 level**, female representation has surpassed parity and "
        "reached 60 per cent.\n\n"
        "Historically, the number of women in leadership roles has increased "
        "steadily. While men continue to represent a larger share overall, "
        "long-term trends point to gradual progress toward greater gender "
        "balance.\n\n"
        "Regionally, leadership representation has diversified, with gradual "
        "increases across Africa, Asia and the Pacific, Eastern Europe, and "
        "Latin America and the Caribbean."
    )

    rows = [
        ["snapshot_month", month],
        ["last_updated", today],
        ["overview", overview],
        ["leadership_on_the_move", leadership_on_move],
        ["monthly_highlight", highlight],
        ["note_characteristics", note_characteristics],
        ["note_trends", note_trends],
        # KPIs match the PPT Slide 1 figures verbatim (these are the
        # February 2026 snapshot values). Recompute from the Google Sheet
        # by leaving these cells blank.
        ["kpi_total_leaders", 27],
        ["kpi_pct_female", 44],
        ["kpi_pct_underrepresented", 60],
        ["kpi_deputy_hcs", 2],
        ["kpi_countries", 25],
    ]
    return header, rows


# ── Orchestration ────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, help="Path to HC trends report XLSX")
    ap.add_argument("--out",  default="data/initial", help="Output directory for CSVs")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        sys.stderr.write(f"ERROR: XLSX not found at {xlsx}\n")
        return 2
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    print(f"Reading {xlsx.name}…")
    wb = load_workbook(xlsx, data_only=True)

    # Identify the sheet that holds the current roster. Prefer the "- LF"
    # (latest filter) tab over the raw report sheet.
    roster_name = next((n for n in wb.sheetnames if n.lower().endswith("- lf")), None)
    if roster_name is None:
        roster_name = next((n for n in wb.sheetnames if "lf" in n.lower()), None)
    if roster_name is None:
        # Fall back to the current year's master report
        roster_name = next((n for n in wb.sheetnames if "2025" in n), wb.sheetnames[0])
    print(f"  current roster → '{roster_name}'")
    roster = sheet_records(wb[roster_name], header_signals=["full name", "fullname"])

    # Keep only rows that look like actual people.
    def looks_like_person(r: Dict) -> bool:
        name = (r.get("full name") or "").strip()
        if not name:
            return False
        lower = name.lower()
        # Filter common non-person rows that sometimes end up in filter views
        junk = ("total", "grand total", "subtotal", "confidential", "copyright",
                "©", "all rights reserved", "source:", "note:")
        if any(j in lower for j in junk):
            return False
        # A real person row usually has either a gender OR a position title
        if not (r.get("gender") or r.get("position title")):
            return False
        return True

    active = [r for r in roster if looks_like_person(r)]
    print(f"  {len(active)} active leaders detected")

    # Pivoted yearly sheets (with title rows offset).
    gender_name = next((n for n in wb.sheetnames if n.lower().strip() == "gender"), None)
    region_name = next((n for n in wb.sheetnames if n.lower().strip() == "region"), None)
    gender_rows = sheet_records(wb[gender_name], header_signals=["row labels", "year"]) if gender_name else []
    region_rows = sheet_records(wb[region_name], header_signals=["row labels", "year"]) if region_name else []
    print(f"  gender trends   → '{gender_name}' ({len(gender_rows)} rows)")
    print(f"  region trends   → '{region_name}' ({len(region_rows)} rows)")

    print(f"\nWriting CSVs → {out}/")
    builders = [
        ("meta.csv",             build_meta(active)),
        ("leaders.csv",          build_leaders(active)),
        ("contacts.csv",         build_contacts(active)),
        ("map_countries.csv",    build_map_countries(active)),
        # Page 3 — exact PPT figures (not derived from the xlsx).
        ("roles_donut.csv",      build_roles_donut_from_ppt(active)),
        ("agency_donut.csv",     build_agency_donut_from_ppt(active)),
        ("country_by_grade.csv", build_country_by_grade(active)),
        ("gender_by_grade.csv",  build_gender_by_grade(active)),
        # Page 4 — historical trends from the xlsx Gender / Region pivots.
        ("gender_trends.csv",    build_gender_trends(gender_rows)),
        ("region_trends.csv",    build_region_trends(region_rows)),
    ]
    for name, (header, rows) in builders:
        write_csv(out / name, header, rows)

    print("\nDone. Next steps:")
    print("  1. Paste each CSV into the matching Google Sheet tab (or leave them here for local dev).")
    print("  2. Open index.html locally: `cd HL_dashboard_BDU && python3 -m http.server 8000`")
    print("  3. Visit http://localhost:8000")
    return 0


if __name__ == "__main__":
    sys.exit(main())
