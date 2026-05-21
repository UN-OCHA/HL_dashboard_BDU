#!/usr/bin/env python3
"""
enrich_tab9_from_xlsx.py
========================

Populate the empty columns in Tab 9 (Leaders roster) of the HL Dashboard
master Google Sheet by looking up each leader in Luiza Fernandes's
"HC trends report" xlsx export.

Cohort policy
-------------
Tab 9 is the canonical roster (29 rows). The xlsx (~62 "currently in
post" + ~370 historical assignments) is NOT a roster — it is a lookup
source for per-leader attributes (gender / nationality / WEOG / agency /
grade). We never add new names to Tab 9 from the xlsx; we only fill in
the empty columns for names already there.

For each Tab 9 leader, we try (in order):
  1. exact name match in xlsx current-in-post rows (End of Assignment empty)
  2. normalised name match (accents stripped, parens stripped, case-fold)
     in current-in-post rows
  3. fall back to historical rows (covers leaders the xlsx flagged as
     ended even though they're still on Valijon's roster)

If nothing matches after step 3, the leader's row is left empty and the
report flags it for manual review.

Usage
-----
    export HL_SHEET_URL="https://script.google.com/.../exec"
    export HL_SHEET_TOKEN="<shared secret>"
    /tmp/hl-xlsx-venv/bin/python3 scripts/enrich_tab9_from_xlsx.py \
        --xlsx "../data/HC trends report 2025.xlsx" \
        --tab  "9. Leaders (roster)" \
        [--dry-run]

`--dry-run` reads everything, prints the diff, but does NOT push.
Always run dry-run first when re-running against a new xlsx.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from pathlib import Path
from typing import Iterable

# Third-party — install in the helper venv:
#   python3 -m venv /tmp/hl-xlsx-venv
#   /tmp/hl-xlsx-venv/bin/pip install openpyxl
from openpyxl import load_workbook


# ── Column mapping (Tab 9 is row 1 = banner, row 2 = header, row 3+ data)
TAB9_HEADER = [
    "country", "duty_station", "name", "position",
    "hat3", "gender", "nationality", "weog", "agency", "eod", "eoa",
]
WRITE_COLS = ["hat3", "gender", "nationality", "weog", "agency", "eod", "eoa"]
NAME_COL_IDX = TAB9_HEADER.index("name")          # 2 (0-indexed)
DATA_START_ROW = 3                                # 1-indexed (banner=1, header=2)

# ── xlsx columns (sheet "HC trends report 01 2025"). 0-indexed.
XLSX_HEADER_ROW = 10                              # 0-indexed → row 11 in Excel
XLSX_COL = {
    "name":        1,   # "Full Name"
    "gender":      4,   # "Gender"
    "weog":        5,   # "WEOG/Non-WEOG"
    "nationality": 6,   # "Nationality"
    "agency":      7,   # "Agency of Origin Short Name"
    "position":    8,   # "Position Title"
    "hat3":        9,   # "Hat 3"
    "eod":        10,   # "Entrance on Duty"
    "eoa":        11,   # "End of Assignment"
}

# Junk-row signals — Salesforce footer leaks into the data. Drop them.
JUNK_CONTAINS = ("Copyright", "Confidential", "Salesforce", "All rights")


def normalise_name(s: str) -> str:
    """Strip accents, parens, hyphens, extra whitespace; case-fold.

    Examples:
        "François Batalingaya"     → "francois batalingaya"
        "Anita (Kiki) Gbeho"       → "anita gbeho"
        "Mireia Villar-Forner"     → "mireia villar forner"
        "  Indrika  Ratwatte "     → "indrika ratwatte"
    """
    if not s:
        return ""
    # Decompose accents → ASCII
    s = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(s))
        if not unicodedata.combining(ch)
    )
    # Strip parenthesised middle-names / nicknames ("Anita (Kiki) Gbeho")
    s = re.sub(r"\s*\([^)]*\)\s*", " ", s)
    # Hyphens → spaces ("Villar-Forner" matches "Villar Forner")
    s = s.replace("-", " ")
    # Collapse whitespace, lowercase
    return re.sub(r"\s+", " ", s).strip().lower()


def first_last(s: str) -> str:
    """Return "first last" — drops middle words. Used as a last-resort match.

    Examples:
        "Mohamed Malick Fall"      → "mohamed fall"
        "Anita Kiki Gbeho"         → "anita gbeho"
        "Indrika Ratwatte"         → "indrika ratwatte"  (already 2 words)
    """
    parts = normalise_name(s).split()
    if len(parts) < 2:
        return ""
    return f"{parts[0]} {parts[-1]}"


def fmt_date(v) -> str:
    """xlsx dates come back as datetime or string. Render YYYY-MM-DD."""
    if v is None or v == "":
        return ""
    try:
        return v.strftime("%Y-%m-%d")
    except AttributeError:
        return str(v).strip()


# ── Sheets API client (POST + JSON, follows the 302 redirect chain
#    that curl can't handle). Mirrors scripts/sheet_call.py.
def api_call(action: str, **kwargs) -> dict:
    url = os.environ["HL_SHEET_URL"]
    token = os.environ["HL_SHEET_TOKEN"]
    body = {"token": token, "action": action, **kwargs}
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        return {"ok": False, "http_status": e.code, "body": e.read().decode("utf-8", errors="replace")}


def load_xlsx_rows(path: Path) -> tuple[list, list]:
    """Return (current_rows, historical_rows) from the trends xlsx.

    "Current" = End of Assignment is empty.
    "Historical" = End of Assignment is set.

    Both lists exclude junk rows (Salesforce footer text).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["HC trends report 01 2025"]
    rows = list(ws.iter_rows(values_only=True))
    data = rows[XLSX_HEADER_ROW + 1:]  # everything after the header row

    current, historical = [], []
    for r in data:
        name = r[XLSX_COL["name"]]
        if not name:
            continue
        if any(j in str(name) for j in JUNK_CONTAINS):
            continue
        (current if r[XLSX_COL["eoa"]] is None else historical).append(r)
    return current, historical


def find_match(target_name: str, current: list, historical: list):
    """Return the best xlsx row for a Tab 9 leader, or None.

    Match tiers (try in order, prefer earliest tier):
      1. exact      — raw string equality
      2. normalised — accents / parens / hyphens / case folded
      3. first_last — first word + last word only (catches middle-name
                      drift: "Mohamed Fall" ↔ "Mohamed Malick Fall")

    Each tier is tried first against current-in-post rows, then against
    historical. For historical hits with multiple candidates, prefer
    the most-recent assignment (latest Entrance on Duty).
    """
    target_norm = normalise_name(target_name)
    target_fl   = first_last(target_name)

    def search(rows, tier: str):
        out = []
        for r in rows:
            xn = r[XLSX_COL["name"]]
            if tier == "exact" and xn == target_name:
                out.append(r)
            elif tier == "normalised" and normalise_name(xn) == target_norm:
                out.append(r)
            elif tier == "first_last" and target_fl and first_last(xn) == target_fl:
                out.append(r)
        return out

    for tier in ("exact", "normalised", "first_last"):
        for pool, label in [(current, "current"), (historical, "historical")]:
            hits = search(pool, tier)
            if hits:
                # For historical hits prefer the most-recent assignment
                if label == "historical" and len(hits) > 1:
                    hits.sort(
                        key=lambda r: (r[XLSX_COL["eod"]] or 0),
                        reverse=True,
                    )
                return hits[0], label, tier
    return None, None, None


def build_values_row(xlsx_row) -> list[str]:
    """Pull the 7 enrichment columns out of an xlsx row, in WRITE_COLS order."""
    if xlsx_row is None:
        return [""] * len(WRITE_COLS)
    return [
        str(xlsx_row[XLSX_COL["hat3"]] or "").strip(),
        str(xlsx_row[XLSX_COL["gender"]] or "").strip(),
        str(xlsx_row[XLSX_COL["nationality"]] or "").strip(),
        str(xlsx_row[XLSX_COL["weog"]] or "").strip(),
        str(xlsx_row[XLSX_COL["agency"]] or "").strip(),
        fmt_date(xlsx_row[XLSX_COL["eod"]]),
        fmt_date(xlsx_row[XLSX_COL["eoa"]]),
    ]


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the HC trends report xlsx")
    ap.add_argument("--tab",  default="9. Leaders (roster)")
    ap.add_argument("--dry-run", action="store_true", help="Read + report, don't push")
    args = ap.parse_args(argv[1:])

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"ERR: xlsx not found at {xlsx_path}", file=sys.stderr)
        return 1

    print(f"→ Reading xlsx: {xlsx_path}")
    current, historical = load_xlsx_rows(xlsx_path)
    print(f"  · {len(current)} current-in-post rows, {len(historical)} historical rows")

    print(f"→ Reading Tab 9 from master sheet…")
    resp = api_call("read", tab=args.tab)
    if not resp.get("ok"):
        print(f"ERR: Sheets API read failed: {resp}", file=sys.stderr)
        return 1
    sheet_rows = resp["rows"]
    leader_rows = [
        (i, r) for i, r in enumerate(sheet_rows[DATA_START_ROW - 1:], start=DATA_START_ROW)
        if r and r[NAME_COL_IDX]
    ]
    print(f"  · {len(leader_rows)} leader rows in Tab 9")

    # ── Match each Tab 9 leader against the xlsx
    enriched_values = []     # 29 × 7 values to push
    report = []
    for sheet_row_num, row in leader_rows:
        name = row[NAME_COL_IDX]
        match, source, exact = find_match(name, current, historical)
        values = build_values_row(match)
        enriched_values.append(values)

        # Report status — `exact` is now the tier name ("exact" / "normalised"
        # / "first_last") returned by find_match.
        status = "NO MATCH" if match is None else f"{source}/{exact}"
        report.append({"row": sheet_row_num, "name": name, "status": status, "values": values})

    # ── Print the report (always)
    print(f"\n=== Enrichment report ({len(report)} leaders) ===")
    print(f"  {'Row':>3}  {'Name':<32s}  {'Match':<22s}  Sample fields")
    matched, partial, missed = 0, 0, 0
    for r in report:
        sample = f"gender={r['values'][1] or '—'}  agency={r['values'][4] or '—'}"
        print(f"  {r['row']:3d}  {r['name'][:32]:<32s}  {r['status']:<22s}  {sample}")
        if r["status"] == "NO MATCH":
            missed += 1
        elif any(v == "" for v in r["values"][:5]):  # hat3/gender/nat/weog/agency
            partial += 1
        else:
            matched += 1

    print(f"\n  Fully matched   : {matched}")
    print(f"  Partial         : {partial}  (matched but some xlsx fields were empty)")
    print(f"  Unmatched       : {missed}  (left blank in Tab 9, flag to Valijon)")

    # ── Push via update_range (cols 5–11 = E–K, rows DATA_START_ROW … N)
    last_row = DATA_START_ROW + len(leader_rows) - 1
    target_range = f"E{DATA_START_ROW}:K{last_row}"
    print(f"\n→ Target range: {args.tab} ! {target_range}")

    if args.dry_run:
        print("  · DRY RUN — not pushing. Re-run without --dry-run to write.")
        return 0

    print("  · Pushing to sheet…")
    resp = api_call(
        "update_range",
        tab=args.tab,
        range=target_range,
        values=enriched_values,
    )
    if not resp.get("ok"):
        print(f"ERR: Sheets API write failed: {resp}", file=sys.stderr)
        return 1
    print(f"  ✓ {resp.get('updated', target_range)}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
